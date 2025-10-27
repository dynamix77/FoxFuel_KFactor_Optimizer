"""
Output generation module for the FoxFuel K-Factor Optimizer.
Generates Apply_K_ThisWeek.csv and K_Review_Queue.xlsx files.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .config import *
from .logger import get_logger

logger = get_logger()

class OutputsWriter:
    """Handles generation of output files for Ignite import and manual review."""
    
    def __init__(self, output_dir: str = OUTPUT_DIR):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def write_apply_k_this_week(self, auto_apply_data: pd.DataFrame) -> Path:
        """
        Generate Apply_K_ThisWeek.csv for Ignite import.
        
        Args:
            auto_apply_data: DataFrame with customers eligible for auto-apply
            
        Returns:
            Path to the generated CSV file
        """
        logger.info("Generating Apply_K_ThisWeek.csv...")
        
        if auto_apply_data is None or len(auto_apply_data) == 0:
            logger.warning("No customers eligible for auto-apply")
            # Create empty file
            output_file = self.output_dir / "Apply_K_ThisWeek.csv"
            pd.DataFrame(columns=['Customer Number', 'K Factor - Winter']).to_csv(output_file, index=False)
            return output_file
        
        # Create the import file (focus on Winter K-factor updates)
        import_data = auto_apply_data[['Customer Number', 'Proposed K Factor']].copy()
        import_data.columns = ['Customer Number', 'K Factor - Winter']
        
        # Sort by customer number
        import_data = import_data.sort_values('Customer Number')
        
        # Write to CSV
        output_file = self.output_dir / "Apply_K_ThisWeek.csv"
        import_data.to_csv(output_file, index=False)
        
        logger.info(f"Generated Apply_K_ThisWeek.csv with {len(import_data)} customers")
        return output_file
    
    def write_k_review_queue(self, governed_data: pd.DataFrame) -> Path:
        """
        Generate K_Review_Queue.xlsx with comprehensive review information.
        
        Args:
            governed_data: DataFrame with all governed K-factors
            
        Returns:
            Path to the generated Excel file
        """
        logger.info("Generating K_Review_Queue.xlsx...")
        
        if governed_data is None or len(governed_data) == 0:
            logger.warning("No governed data provided")
            # Create empty workbook
            output_file = self.output_dir / "K_Review_Queue.xlsx"
            wb = Workbook()
            try:
                wb.save(output_file)
                return output_file
            except PermissionError:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_file = self.output_dir / f"K_Review_Queue_{timestamp}.xlsx"
                wb.save(backup_file)
                logger.warning(f"Original file locked, saved as: {backup_file.name}")
                return backup_file
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create main review sheet
        self._create_review_sheet(wb, governed_data)
        
        # Create summary sheet
        self._create_summary_sheet(wb, governed_data)
        
        # Create auto-apply sheet
        auto_apply_data = self._filter_auto_apply(governed_data)
        if len(auto_apply_data) > 0:
            self._create_auto_apply_sheet(wb, auto_apply_data)
        
        # Save workbook with error handling
        output_file = self.output_dir / "K_Review_Queue.xlsx"
        try:
            wb.save(output_file)
            logger.info(f"Generated K_Review_Queue.xlsx with {len(governed_data)} customers")
            return output_file
        except PermissionError:
            # Try with timestamp if file is locked
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = self.output_dir / f"K_Review_Queue_{timestamp}.xlsx"
            wb.save(backup_file)
            logger.warning(f"Original file locked, saved as: {backup_file.name}")
            return backup_file
    
    def _create_review_sheet(self, wb: Workbook, governed_data: pd.DataFrame):
        """Create the main review sheet with conditional formatting."""
        
        ws = wb.create_sheet("K Review Queue")
        
        # Prepare data for Excel
        review_data = governed_data[[
            'Customer Number', 'K Factor - Winter', 'Weighted K Factor', 'Proposed K Factor',
            'Final Variance Percent', 'Final Status', 'Confidence', 'Interval Count',
            'Total Gallons', 'Total Degree Days'
        ]].copy()
        
        # Add run-out risk flag
        review_data['Run-Out Risk'] = review_data['Final Variance Percent'] < -20
        
        # Sort by priority
        priority_order = {
            'HIGH_VARIANCE': 1, 'LOW_CONFIDENCE': 2, 'INSUFFICIENT_DATA': 3,
            'CAPPED_INCREASE': 4, 'CAPPED_DECREASE': 4, 'CAPPED_UPPER_BOUND': 4, 'CAPPED_LOWER_BOUND': 4,
            'APPROVED': 5
        }
        review_data['Priority'] = review_data['Final Status'].map(priority_order)
        review_data = review_data.sort_values(['Priority', 'Final Variance Percent'], ascending=[True, False])
        
        # Add headers (Winter/Summer K-factor focus)
        headers = [
            'Customer Number', 'Current Winter K Factor', 'Calculated K', 'Proposed K',
            'Variance %', 'Status', 'Confidence', 'Intervals', 'Total Gallons',
            'Total Degree Days', 'Run-Out Risk'
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, (_, row) in enumerate(review_data.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row['Customer Number'])
            ws.cell(row=row_idx, column=2, value=row['K Factor - Winter'])
            ws.cell(row=row_idx, column=3, value=row['Weighted K Factor'])
            ws.cell(row=row_idx, column=4, value=row['Proposed K Factor'])
            ws.cell(row=row_idx, column=5, value=row['Final Variance Percent'])
            ws.cell(row=row_idx, column=6, value=row['Final Status'])
            ws.cell(row=row_idx, column=7, value=row['Confidence'])
            ws.cell(row=row_idx, column=8, value=row['Interval Count'])
            ws.cell(row=row_idx, column=9, value=row['Total Gallons'])
            ws.cell(row=row_idx, column=10, value=row['Total Degree Days'])
            ws.cell(row=row_idx, column=11, value="YES" if row['Run-Out Risk'] else "NO")
            
            # Apply conditional formatting
            self._apply_row_formatting(ws, row_idx, row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_row_formatting(self, ws, row_idx: int, row_data: pd.Series):
        """Apply conditional formatting to a row based on status and variance."""
        
        # Color coding based on status
        if row_data['Final Status'] == 'HIGH_VARIANCE':
            fill_color = "FFE6E6"  # Light red
        elif row_data['Final Status'] == 'LOW_CONFIDENCE':
            fill_color = "FFF2CC"  # Light yellow
        elif row_data['Final Status'] == 'INSUFFICIENT_DATA':
            fill_color = "E6F3FF"  # Light blue
        elif 'CAPPED' in row_data['Final Status']:
            fill_color = "F0F0F0"  # Light gray
        else:
            fill_color = "E6FFE6"  # Light green
        
        # Apply fill to entire row
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
        
        # Highlight run-out risk
        if row_data['Run-Out Risk']:
            ws.cell(row=row_idx, column=11).font = Font(bold=True, color="FF0000")
    
    def _create_summary_sheet(self, wb: Workbook, governed_data: pd.DataFrame):
        """Create summary statistics sheet."""
        
        ws = wb.create_sheet("Summary")
        
        # Calculate summary statistics
        total_customers = len(governed_data)
        auto_apply_count = len(self._filter_auto_apply(governed_data))
        review_count = total_customers - auto_apply_count
        
        high_variance = len(governed_data[governed_data['Final Status'] == 'HIGH_VARIANCE'])
        low_confidence = len(governed_data[governed_data['Final Status'] == 'LOW_CONFIDENCE'])
        insufficient_data = len(governed_data[governed_data['Final Status'] == 'INSUFFICIENT_DATA'])
        
        avg_variance = governed_data['Final Variance Percent'].mean()
        avg_confidence = governed_data['Confidence'].mean()
        
        # Write summary
        summary_data = [
            ["K-Factor Optimization Summary", ""],
            ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["Total Customers Processed:", total_customers],
            ["Auto-Apply Eligible:", auto_apply_count],
            ["Manual Review Required:", review_count],
            ["", ""],
            ["Status Breakdown:", ""],
            ["High Variance:", high_variance],
            ["Low Confidence:", low_confidence],
            ["Insufficient Data:", insufficient_data],
            ["", ""],
            ["Average Variance %:", f"{avg_variance:.2f}%"],
            ["Average Confidence:", f"{avg_confidence:.3f}"],
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            
            if row_idx == 1:  # Title
                ws.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
            elif label.endswith(":"):  # Labels
                ws.cell(row=row_idx, column=1).font = Font(bold=True)
    
    def _create_auto_apply_sheet(self, wb: Workbook, auto_apply_data: pd.DataFrame):
        """Create auto-apply sheet for easy reference."""
        
        ws = wb.create_sheet("Auto-Apply")
        
        # Write headers
        headers = ['Customer Number', 'Current Winter K Factor', 'Proposed Winter K Factor', 'Variance %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Write data
        for row_idx, (_, row) in enumerate(auto_apply_data.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row['Customer Number'])
            ws.cell(row=row_idx, column=2, value=row['K Factor - Winter'])
            ws.cell(row=row_idx, column=3, value=row['Proposed K Factor'])
            ws.cell(row=row_idx, column=4, value=row['Final Variance Percent'])
    
    def _filter_auto_apply(self, governed_data: pd.DataFrame) -> pd.DataFrame:
        """Filter customers eligible for auto-apply."""
        return governed_data[
            (governed_data['Confidence'] >= CONFIDENCE_THRESHOLD) &
            (governed_data['Interval Count'] >= MIN_INTERVALS) &
            (governed_data['Final Status'].isin(['APPROVED', 'CAPPED_INCREASE', 'CAPPED_DECREASE', 
                                                'CAPPED_UPPER_BOUND', 'CAPPED_LOWER_BOUND']))
        ]
